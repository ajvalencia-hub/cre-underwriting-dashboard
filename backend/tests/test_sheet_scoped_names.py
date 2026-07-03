"""Regression tests for FINDINGS.md M1: worksheet-scoped defined names are
invisible in wb.defined_names (verified against openpyxl 3.1.5), so templates
using them silently lost those mapping candidates and could never resolve
them at generate time. They're now listed qualified as 'Sheet!Name' and
resolve end-to-end through inject_values / read_output_values.
"""

import openpyxl
import pytest
from openpyxl.workbook.defined_name import DefinedName

from app.services import template_service
from app.services.excel_writer import inject_values, read_output_values


@pytest.fixture
def template(tmp_path):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Inputs"
    ws2 = wb.create_sheet("My Model")  # space in the title on purpose
    wb.defined_names["GlobalPrice"] = DefinedName("GlobalPrice", attr_text="Inputs!$A$1")
    ws2.defined_names["LocalCap"] = DefinedName("LocalCap", attr_text="'My Model'!$B$2")
    path = tmp_path / "template.xlsx"
    wb.save(path)
    return path


def test_parse_workbook_lists_sheet_scoped_names_qualified(template):
    parsed = template_service.parse_workbook(template)
    by_name = {nr["name"]: nr for nr in parsed["namedRanges"]}
    assert "GlobalPrice" in by_name
    assert "My Model!LocalCap" in by_name
    assert by_name["My Model!LocalCap"]["sheet"] == "My Model"
    assert by_name["My Model!LocalCap"]["ref"] == "B2"


def test_inject_resolves_sheet_scoped_name(template, tmp_path):
    out = tmp_path / "out.xlsx"
    mappings = {
        "purchasePrice": {"target": "namedRange", "ref": "GlobalPrice"},
        "exitCapRate": {"target": "namedRange", "ref": "My Model!LocalCap"},
    }
    result = inject_values(template, out, mappings, {"purchasePrice": 1000000, "exitCapRate": 0.055})

    assert sorted(result["written"]) == ["exitCapRate", "purchasePrice"]
    assert result["warnings"] == []
    wb = openpyxl.load_workbook(out)
    assert wb["My Model"]["B2"].value == 0.055


def test_read_output_values_resolves_sheet_scoped_name(template, tmp_path):
    out = tmp_path / "out.xlsx"
    mappings = {"exitCapRate": {"target": "namedRange", "ref": "My Model!LocalCap"}}
    inject_values(template, out, mappings, {"exitCapRate": 0.055})
    outputs = read_output_values(out, mappings, ["exitCapRate"])
    assert outputs["exitCapRate"] == 0.055


def test_unknown_sheet_qualifier_skips_with_warning(template, tmp_path):
    out = tmp_path / "out.xlsx"
    mappings = {"x": {"target": "namedRange", "ref": "Nope!LocalCap"}}
    result = inject_values(template, out, mappings, {"x": 1})
    assert result["written"] == []
    assert any("Could not resolve" in w for w in result["warnings"])
