"""Regression tests for FINDINGS.md H4/H5: injection edge cases that crashed
the whole generate flow (multi-cell named ranges, merged-cell targets).
"""

import openpyxl
import pytest
from openpyxl.workbook.defined_name import DefinedName

from app.services.excel_writer import inject_values, read_output_values


@pytest.fixture
def template(tmp_path):
    """A workbook with a multi-cell named range, a scalar named range, and a
    merged region — the exact shapes that used to raise AttributeError."""
    path = tmp_path / "template.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "My Sheet"  # space on purpose
    ws["B1"] = 0
    ws["D1"] = "=B1*2"
    ws.merge_cells("A3:C3")
    ws["A3"] = None
    wb.defined_names.add(DefinedName("ScalarName", attr_text="'My Sheet'!$B$1"))
    wb.defined_names.add(DefinedName("RangeName", attr_text="'My Sheet'!$B$1:$C$2"))
    wb.save(path)
    return path


def test_multi_cell_named_range_is_skipped_with_warning_not_crash(template, tmp_path):
    out = tmp_path / "out.xlsx"
    mappings = {
        "badField": {"target": "namedRange", "ref": "RangeName"},
        "goodField": {"target": "namedRange", "ref": "ScalarName"},
    }
    result = inject_values(template, out, mappings, {"badField": 5, "goodField": 7})

    assert "badField" not in result["written"]
    assert any("multi-cell range" in w and "badField" in w for w in result["warnings"])
    # The rest of the injection still happened:
    assert "goodField" in result["written"]
    wb = openpyxl.load_workbook(out)
    assert wb["My Sheet"]["B1"].value == 7


def test_read_output_values_skips_multi_cell_named_range(template):
    mappings = {"m": {"target": "namedRange", "ref": "RangeName"}}
    assert read_output_values(template, mappings, ["m"]) == {}


def test_merged_non_anchor_cell_writes_to_anchor_not_crash(template, tmp_path):
    out = tmp_path / "out.xlsx"
    # B3 is a non-anchor cell of the merged A3:C3 — pre-fix this raised
    # AttributeError ('MergedCell' value is read-only) and 500ed the request.
    mappings = {"merged": {"target": "cell", "ref": "B3", "sheet": "My Sheet"}}
    result = inject_values(template, out, mappings, {"merged": 42})

    assert "merged" in result["written"]
    wb = openpyxl.load_workbook(out)
    assert wb["My Sheet"]["A3"].value == 42  # landed on the anchor


def test_table_injection_skips_merged_cells_with_warning(template, tmp_path):
    out = tmp_path / "out.xlsx"
    # Anchor the table at A2 with 2 rows: the second row overlaps merged A3:C3.
    mappings = {
        "tbl": {
            "target": "table",
            "anchor": "A2",
            "sheet": "My Sheet",
            "columnOrder": ["a", "b"],
        }
    }
    rows = [{"a": 1, "b": 2}, {"a": 3, "b": 4}]
    result = inject_values(template, out, mappings, {"tbl": rows})

    assert "tbl" in result["written"]
    assert any("merged ranges" in w for w in result["warnings"])
    wb = openpyxl.load_workbook(out)
    ws = wb["My Sheet"]
    assert ws["A2"].value == 1 and ws["B2"].value == 2  # unmerged row written
    # Only the anchor A3 is writable in the merged row; B3 was skipped, and
    # nothing was written to A3 by the anchor-overlap either (A3 is the merge
    # anchor — the loop targets it directly, so it does get row 2's "a"):
    assert ws["A3"].value == 3


def test_read_output_values_reads_merged_anchor_value(template, tmp_path):
    out = tmp_path / "out.xlsx"
    inject_values(template, out, {"m": {"target": "cell", "ref": "A3", "sheet": "My Sheet"}}, {"m": 9})
    # Map the OUTPUT to the non-anchor B3; the readable value lives on A3.
    mappings = {"o": {"target": "cell", "ref": "B3", "sheet": "My Sheet"}}
    assert read_output_values(out, mappings, ["o"]) == {"o": 9}
