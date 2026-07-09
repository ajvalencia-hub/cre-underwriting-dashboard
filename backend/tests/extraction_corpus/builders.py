"""Synthetic extraction fixtures shaped like the files brokers and property
managers actually send, with the hostile formatting the audit found in the
wild baked in on purpose:

- Yardi-Voyager-style rent roll: merged title banner rows above the headers,
  text unit numbers ("A-101"), a literal VACANT resident, a blank vacant row,
  and a mid-table subtotal row.
- Yardi-style T-12: twelve month columns plus a trailing annual total,
  parenthesized-negative strings, subtotal/NOI rows, and lines that match no
  standard category (Pest Control, RUBS Income).
- RealPage-style rent roll: different header vocabulary and an explicit
  status column.
- Broker-OM-style PDF: a two-page ruled rent-roll table with a repeated
  header row (exercises the multi-page merge fix).
"""

import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Table, TableStyle

MONTH_HEADERS = [
    "Jan 2026", "Feb 2026", "Mar 2026", "Apr 2026", "May 2026", "Jun 2026",
    "Jul 2026", "Aug 2026", "Sep 2026", "Oct 2026", "Nov 2026", "Dec 2026",
]


def build_yardi_rent_roll(path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rent Roll"
    ws.merge_cells("A1:H1")
    ws["A1"] = "Maple Court Apartments"
    ws.merge_cells("A2:H2")
    ws["A2"] = "Rent Roll as of 06/30/2026"
    headers = ["Unit", "Unit Type", "SQFT", "Resident", "Market Rent", "Actual Rent", "Move In", "Lease Expiration"]
    ws.append([])
    ws.append(headers)
    rows = [
        ["A-101", "1BR/1BA", 750, "Alice Johnson", 1500, 1450, "05/01/2024", "04/30/2026"],
        ["A-102", "1BR/1BA", 750, "Bob Smith", 1500, 1480, "01/15/2025", "01/14/2027"],
        ["A-103", "1BR/1BA", 750, "VACANT", 1500, None, None, None],
        ["Total 1BR/1BA", None, 2250, None, 4500, 2930, None, None],  # subtotal row
        ["B-201", "2BR/2BA", 1100, "Carol Davis", 2100, 2050, "07/01/2023", "06/30/2026"],
        ["B-202", "2BR/2BA", 1100, "Dan Edwards", 2100, 2080, "03/01/2025", "02/28/2027"],
        ["B-203", "2BR/2BA", 1100, None, 2100, None, None, None],  # blank vacant row
        ["Total 2BR/2BA", None, 3300, None, 6300, 4130, None, None],
    ]
    for row in rows:
        ws.append(row)
    wb.save(path)


def build_yardi_t12(path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "T12"
    ws.merge_cells("A1:N1")
    ws["A1"] = "Maple Court Apartments — Trailing 12 Statement"
    ws.append(["Account"] + MONTH_HEADERS + ["Annual Total"])

    def line(label, monthly, total=None, as_text=False):
        cells = [label]
        for _ in range(12):
            cells.append(f"({abs(monthly):,})" if as_text and monthly < 0 else monthly)
        cells.append(
            f"({abs(monthly) * 12:,})" if as_text and monthly < 0 else (total if total is not None else monthly * 12)
        )
        ws.append(cells)

    line("Gross Potential Rent", 50_000)
    line("Vacancy Loss", -2_500, as_text=True)  # parenthesized-negative strings
    line("Bad Debt", -400, as_text=True)
    line("RUBS Income", 1_200)  # matches no standard alias -> unclassified
    line("Total Income", 48_300)
    line("Real Estate Taxes", 5_000)
    line("Insurance", 1_500)
    line("Utilities", 2_000)
    line("Repairs and Maintenance", 1_800)
    line("Payroll", 2_500)
    line("Management Fees", 1_600)
    line("Office Expense", 700)
    line("Reserve for Replacement", 500)
    line("Pest Control", 250)  # unclassified expense line
    line("Total Operating Expenses", 15_850)
    line("Net Operating Income", 32_450)
    wb.save(path)


def build_realpage_rent_roll(path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Unit Availability"
    headers = ["Bldg/Unit", "Floorplan", "SQFT", "Name", "Lease Rent", "Market + Addl.", "Move-In", "Lease End", "Status"]  # noqa: E501
    ws.append(headers)
    rows = [
        ["01-101", "A1", 720, "Eve Foster", 1395, 1425, "2024-09-01", "2026-08-31", "Occupied"],
        ["01-102", "A1", 720, "Frank Garcia", 1410, 1425, "2025-02-01", "2027-01-31", "Occupied"],
        ["01-103", "A1", 720, None, None, 1425, None, None, "Vacant-Ready"],
        ["02-201", "B2", 1050, "Grace Harris", 1975, 2010, "2023-11-15", "2026-11-14", "Occupied"],
        ["02-202", "B2", 1050, "Henry Irving", 1990, 2010, "2025-05-01", "2026-04-30", "Notice"],
    ]
    for row in rows:
        ws.append(row)
    wb.save(path)


def build_studio_heavy_rent_roll_with_tenant_id(path) -> None:
    """Regression fixture (post-M audit): a studio-majority multifamily roll
    with BOTH a "Tenant ID" and a "Resident Name" column, and hyphenated
    unit-type labels ("1-Bed 1-Bath", "2-Bed 2-Bath") — the exact shape that
    slipped through three separate bugs on a real 64-unit deal:
    (1) _MULTIFAMILY_UNIT_TYPE_RE missed the hyphen, routing the whole roll
        down the commercial-lease path; (2) even fixed, a studio-majority
        mix (studios carry no bed-count digit at all, by design) still fell
        under the >0.5 match-ratio threshold until studios counted too;
        (3) the "tenant" field bound to Tenant ID (a number) instead of
        Resident Name, silently defeating the "VACANT" vacancy marker this
        fixture puts in Resident Name specifically, not Tenant ID."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rent Roll"
    headers = ["Unit", "Unit Type", "Sq Ft", "Tenant ID", "Resident Name", "Lease Start", "Lease End", "Market Rent", "RC (Rent)"]  # noqa: E501
    ws.append(headers)
    rows = [
        ["S-101", "Studio", 480, "1001", "Ada Lovelace", "2024-01-01", "2026-12-31", 1550, 1500],
        ["S-102", "Studio", 480, "1002", "Grace Hopper", "2024-02-01", "2026-01-31", 1550, 1500],
        ["S-103", "Studio", 480, "1003", "Alan Turing", "2024-03-01", "2026-02-28", 1550, 1500],
        ["S-104", "Studio", 480, None, "VACANT", None, None, 1550, 1400],
        ["L-201", "1-Bed 1-Bath", 625, "1004", "Katherine Johnson", "2024-04-01", "2027-03-31", 1776, 1673],
        ["T-301", "2-Bed 2-Bath", 1000, "1005", "Dorothy Vaughan", "2024-05-01", "2027-04-30", 2300, 2225],
    ]
    for row in rows:
        ws.append(row)
    wb.save(path)


def build_property_management_t12_with_banner_row(path) -> None:
    """Regression fixture (post-M audit): an AppFolio-style cash-basis T-12
    with (a) several single-cell metadata/title rows before the real header
    row, (b) a MERGED single-row banner ("Owner's Actuals", A1:P1-style)
    immediately above the real month-header row — scoring higher than a
    bare title row but, before the parse_numeric fix below, tying the real
    header row's score because parse_numeric("JUN 25") used to silently
    return 25.0 (stripping the letters) instead of None, making the
    header-row-guesser's text/number cell scoring blind to every month
    header — and (c) account-code-prefixed line-item labels ("411010 Rental
    Income"), which must still classify correctly once the real header row
    is found."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Profit & Loss 12 Month Recap"])
    ws.append(["Monthly recap from 06/01/25 to 05/31/26"])
    ws.append(["Cash Basis"])
    ws.append(["Property: TEST HOLDINGS"])
    ws.append([])
    banner_row = 6
    ws.cell(row=banner_row, column=1, value="Owner's Actuals")
    ws.merge_cells(start_row=banner_row, start_column=1, end_row=banner_row, end_column=13)
    ws.append([""] + MONTH_HEADERS + ["TOTAL"])
    ws.append(["Income"])
    ws.append(["411010 Rental Income"] + [50_000] * 12 + [600_000])
    ws.append(["421120 Water Charge collected"] + [1_000] * 12 + [12_000])
    ws.append(["Total Income"] + [51_000] * 12 + [612_000])
    ws.append(["Expense"])
    ws.append(["501510 Exp:Prop-Taxes-Paid"] + [2_000] * 12 + [24_000])
    ws.append(["502832 Electricity - Common Area"] + [500] * 12 + [6_000])
    ws.append(["Total Expense"] + [2_500] * 12 + [30_000])
    ws.append(["NET INCOME"] + [48_500] * 12 + [582_000])
    wb.save(path)


def build_commercial_rent_roll(path) -> None:
    """Office/retail-style commercial roll: suite/tenant/SF/monthly rent/
    lease type/dates, mixed date formats, one vacant suite. Feeds the H1
    lease-proposal path (rent converts to $psf/yr; lease types map to
    recovery structures; the vacant suite is skipped with a warning)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rent Roll"
    headers = ["Suite", "Tenant", "SF", "Monthly Rent", "Lease Type", "Lease Start", "Lease Expiration"]
    ws.append(headers)
    rows = [
        ["100", "Blue Bagel LLC", 2400, 6000, "NNN", "01/01/2024", "12/31/2028"],
        ["110", "Verde Yoga", 1800, 3900, "Gross", "06/01/2025", "05/31/2030"],
        ["120", "Corner Dental", 3000, 8250, "Modified Gross", "2023-03-01", "2033-02-28"],
        ["130", "VACANT", 1200, None, None, None, None],
    ]
    for row in rows:
        ws.append(row)
    wb.save(path)


def build_costar_rent_roll(path) -> None:
    """CoStar-export-style commercial roll (I9): their column vocabulary —
    Tenant Name / Suite / Floor / SF Leased / Lease Commencement / Lease
    Expiration / Annual Rent / Rent/SF/Yr / Lease Type. No monthly rent
    column at all: monthly derives from the annual figures."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lease Comps Export"
    headers = [
        "Tenant Name", "Suite", "Floor", "SF Leased", "Lease Commencement",
        "Lease Expiration", "Annual Rent", "Rent/SF/Yr", "Lease Type",
    ]
    ws.append(headers)
    rows = [
        ["Meridian Legal Group", "400", "4", 5200, "03/01/2024", "02/28/2031",
         197600, 38.00, "Full Service"],
        ["Atlas Wealth Advisors", "410", "4", 3100, "07/01/2025", "06/30/2032",
         111600, 36.00, "Full Service"],
        ["Pearl Diagnostics", "200", "2", 6800, "01/01/2023", "12/31/2029",
         231200, 34.00, "NNN"],
        ["Harbor Title Co", "210", "2", 2400, "09/01/2024", "08/31/2029",
         None, 33.50, "Modified Gross"],  # annual absent -> derives from $/SF
    ]
    for row in rows:
        ws.append(row)
    wb.save(path)


def build_costar_rent_roll_hostile(path) -> None:
    """The hostile CoStar variant (I9): month-year-only dates, an MTM term,
    a combined suite range, and a rent column mixing MONTHLY and ANNUAL
    magnitudes (the annual one detected by the magnitude heuristic)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rent Roll"
    headers = ["Suite", "Tenant Name", "SF Leased", "Rent", "Lease Type",
               "Lease Commencement", "Lease Expiration"]
    ws.append(headers)
    rows = [
        # monthly-magnitude rent: 4,500/mo on 1,800sf -> $30/SF/yr, plausible
        ["150", "Cobalt Coffee", 1800, 4500, "NNN", "05/01/2024", "Jun 2027"],
        # ANNUAL-magnitude rent in the same column: 87,500 on 2,500sf read as
        # monthly implies $420/SF/yr -> reinterpreted as annual ($35/SF)
        ["Suites 100-102", "Vantage Media", 2500, 87500, "Full Service", "01/2024", "06/2028"],
        # MTM tenant: no expiry proposed, warned
        ["160", "Quick Print Kiosk", 600, 1500, "Gross", "02/01/2020", "MTM"],
    ]
    for row in rows:
        ws.append(row)
    wb.save(path)


def build_stacking_plan_pdf(path) -> None:
    """Broker-OM stacking-plan table (I9): Floor / Suite / Tenant / SF /
    Expiry with SPARSE rents (a Rent PSF column filled on some rows only).
    Occupied no-rent rows must survive as $0 proposals, not vanish."""
    styles = getSampleStyleSheet()
    header = ["Floor", "Suite", "Tenant", "SF", "Rent PSF", "Expiration"]
    rows = [header] + [
        ["5", "500", "Summit Engineering", "8,000", "31.50", "12/31/2030"],
        ["4", "400", "Beacon Insurance", "8,000", "", "06/30/2028"],
        ["3", "300", "Cedar & Field LLP", "7,500", "29.00", "Mar 2029"],
        ["2", "200", "", "7,500", "", ""],  # vacant floor
        ["1", "100", "Lobby Retail Partners", "4,000", "42.00", "09/30/2027"],
    ]
    grid_style = TableStyle(
        [
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ]
    )
    doc = SimpleDocTemplate(str(path), pagesize=letter)
    doc.build(
        [
            Paragraph("Offering Memorandum — Meridian Tower", styles["Title"]),
            Paragraph("Stacking plan as of June 2026.", styles["Normal"]),
            Table(rows, style=grid_style),
        ]
    )


def build_combined_rent_roll_and_income_statement(path) -> None:
    """A small broker workbook shaped exactly like a real-world failure mode
    (traced from an actual deal package): ONE sheet stacking a rent roll —
    with NO Unit Type column at all, "Unit N" ids, and vacant units marked
    only in the unit LABEL ("Unit 3 - Vacant") while the tenant column still
    carries a generic "Residential" placeholder — directly above a simple
    two-column "label: value" income statement with a CURRENT IN-PLACE
    section and a PRO-FORMA section repeating several of the same expense
    labels at different amounts, plus an "Asking Price:" aside sharing a row
    with an unrelated expense line. Exercises: rent-roll table-boundary
    detection (the income-statement rows below must not become phantom
    units), vacant-by-label inference, the no-unit-type multifamily
    fallback, SF-based unit-mix grouping, and the label/value
    operating-statement parser's section-priority + same-bucket summing."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Unit #", "Tenant", "SF", "Monthly Rent"])
    ws.append(["Unit 1", "Residential", 400, 1200])
    ws.append(["Unit 2", "Residential", 400, 1250])
    ws.append(["Unit 3 - Vacant", "Residential", 400, None])
    ws.append(["Unit 4", "Residential", 600, 1600])
    ws.append(["TOTAL:", None, 1800, 4050])
    ws.append(["AVERAGE:", None, 450, None])
    ws.append([])
    ws.append(["Notes"])
    ws.append(["Figures are approximate."])
    ws.append([])
    ws.append(["CURRENT IN-PLACE"])
    ws.append(["Gross Annual Income", 61200])
    ws.append(["Property Taxes", 8000, None, "Asking Price:", 900000])
    ws.append(["Electric", 600])
    ws.append(["Water/Sewer", 900])
    ws.append(["IN-PLACE NET OPERATING INCOME:", 43700])
    ws.append([])
    ws.append(["PRO-FORMA"])
    ws.append(["Gross Annual Income", 76800])
    ws.append(["Property Taxes", 8200])
    ws.append(["Electric", 600])
    ws.append(["Water/Sewer", 900])
    ws.append(["PRO FORMA NET OPERATING INCOME: ", 58900])
    ws.append(["PRO-FORMA NOI @ 100% OCCUPANCY w/ 3rd Party Management: ", 61000])
    wb.save(path)


def build_marketing_om_without_literal_phrase_pdf(path) -> None:
    """Real-world OM failure mode (traced from an actual broker package):
    a multi-page marketing deck that never once uses the phrase "offering
    memorandum" anywhere in its text — the cover/highlights/zoning/photo
    pages are the ONLY content, no financials at all. Generic real-estate
    boilerplate ("Unit Count: 12", "Bldg Area: 5,677 SF") on the fact-sheet
    page is exactly the kind of scattered wording that used to outscore the
    real OM-specific vocabulary and get this misclassified as a rent roll."""
    styles = getSampleStyleSheet()
    footer = "Jane Broker, Senior Commercial Advisor  |  555-0100  |  Acme Commercial Advisors LLC"

    def page(*paragraphs):
        return [Paragraph(p, styles["Normal"]) for p in paragraphs] + [
            Paragraph(footer, styles["Normal"]),
            PageBreak(),
        ]

    flow = []
    flow += page("Maple Court Apartments", "123 Maple St, Springfield")
    flow += page(
        "Investment Highlights",
        "Stabilized 12-Unit Multifamily Asset offering reliable cash flow.",
        "Unit Count: 12", "Bldg Area: 5,677 SF", "Lot Size: 10,395 SF",
    )
    flow += page(
        "Zoning", "Subject Zoning: T5-R", "Max. Density: 15 units",
        "Max. Height: 5 stories", "Allowable Uses: Multi-family",
    )
    flow += page("Bird's Eye View", "Unit Mix diagram — see site plan.")
    flow += page("Building Photos")
    flow += page("Exterior Photos")
    flow += page("Interior Photos")
    flow += page("Neighborhood Map", "Prime location near transit and retail.")
    flow += page(footer)  # 9th page, keeps page_count comfortably >= 8

    doc = SimpleDocTemplate(str(path), pagesize=letter)
    doc.build(flow)


def build_broker_om_pdf(path) -> None:
    styles = getSampleStyleSheet()
    header = ["Unit", "Tenant", "SF", "Rent"]
    page1_rows = [header] + [
        ["101", "Ivy Café LLC", "1,200", "3,600"],
        ["102", "Jasper Books", "950", "2,375"],
        ["103", "Kite Fitness", "2,100", "5,250"],
    ]
    page2_rows = [header] + [  # repeated header on the continuation page
        ["104", "Luna Salon", "800", "2,200"],
        ["105", "", "1,500", ""],  # vacant suite
    ]
    grid_style = TableStyle(
        [
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ]
    )
    doc = SimpleDocTemplate(str(path), pagesize=letter)
    doc.build(
        [
            Paragraph("Offering Memorandum — Shoppes at Kite Hill", styles["Title"]),
            Paragraph("Confidential. Rent roll as of June 2026.", styles["Normal"]),
            Table(page1_rows, style=grid_style),
            PageBreak(),
            Table(page2_rows, style=grid_style),
        ]
    )
