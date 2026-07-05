"""H11: native Excel model export — structural checks that don't need
LibreOffice (the value parity lives in tests/parity/run.py as
export_analytic_acquisition / export_amortizing_growth)."""

import json
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app.main import app
from app.services import excel_model_export
from app.services.excel_model_export import (
    OUTPUT_CELLS,
    UnsupportedModelFeatures,
    build_model_workbook,
)
from app.services.proforma import engine

FIXTURES = Path(__file__).parent / "fixtures"


@pytest.fixture
def analytic() -> dict:
    return json.loads((FIXTURES / "analytic_acquisition.json").read_text())


def _workbook(inputs) -> openpyxl.Workbook:
    content, _ = build_model_workbook(inputs)
    return openpyxl.load_workbook(BytesIO(content))


def test_unsupported_features_refuse_with_the_full_list(analytic):
    hostile = {
        **analytic,
        "waterfallTiers": [{"hurdlePct": 0.08}],
        "irrConvention": "xirr",
        "useReassessedTaxes": True,
        "commercialLeases": [{
            "tenant": "T", "sf": 1000, "startDate": "2026-01-01",
            "endDate": "2036-01-01", "baseRentPsfAnnual": 30,
        }],
    }
    with pytest.raises(UnsupportedModelFeatures) as excinfo:
        build_model_workbook(hostile)
    text = str(excinfo.value)
    for fragment in ("lease", "waterfall", "XIRR", "reassessed"):
        assert fragment in text, fragment


def test_development_sold_before_stabilization_refuses():
    dev = {
        "dealType": "development",
        "landCost": 1_000_000, "hardCosts": 5_000_000, "softCosts": 1_000_000,
        "constructionMonths": 20, "leaseUpMonths": 12,  # stabilizes month 33
        "grossPotentialRent": 800_000,
        "holdPeriodYears": 2,  # exits month 24 — before stabilization
        "exitCapRatePct": 0.06,
    }
    with pytest.raises(UnsupportedModelFeatures) as excinfo:
        build_model_workbook(dev)
    assert "sold before stabilization" in str(excinfo.value)


def _input_cell(wb, label_fragment: str):
    ws = wb["Inputs"]
    for row in ws.iter_rows(min_col=1, max_col=2):
        if row[0].value and label_fragment in str(row[0].value):
            return row[1].value
    raise AssertionError(f"No Inputs row labeled like {label_fragment!r}")


def test_workbook_is_formula_live_with_sized_loan(analytic):
    wb = _workbook(analytic)
    assert wb.calculation.fullCalcOnLoad
    out = wb["Outputs"]
    assert out[OUTPUT_CELLS["leveredIrr"]].value.startswith("=(1+IRR(")
    model = wb["Model"]
    assert str(model["D2"].value).startswith("=IF($B2<1,0,Inputs!")  # GPR is a formula
    assert model["K2"].value == "=H2-I2-J2"  # NOI identity
    # 60 hold months + 12 forward NOI months (NOI = column K)
    assert model.cell(row=73, column=11).value is not None
    assert model.cell(row=74, column=2).value is None
    # The Debt tab ties to the Model schedule by reference.
    assert wb["Debt"]["B2"].value == "=Model!$L$2"

    # Loan cell carries the engine's sized amount as a VALUE.
    loan = engine.compute(analytic)["debt"]["loanAmount"]
    assert _input_cell(wb, "Loan amount (sized by app)") == pytest.approx(loan)


def test_expenses_block_resolves_bases_and_annotates_recoverable(analytic):
    detail = {
        **analytic,
        "unitMix": [{"unitType": "1BR", "unitCount": 50, "inPlaceRent": 1_500}],
        "grossPotentialRent": 0,
        "opexLineItems": [
            {"category": "taxes", "amount": 30_000, "basis": "annual_total", "recoverable": "yes"},
            {"category": "utilities", "amount": 400, "basis": "per_unit", "recoverable": "no"},
            {"category": "management_fee", "amount": 0.03, "basis": "pct_of_egi"},
        ],
        "realEstateTaxes": 0,
    }
    wb = _workbook(detail)
    ws = wb["Inputs"]
    rows = {
        str(row[0].value): row
        for row in ws.iter_rows(min_col=1, max_col=6)
        if row[0].value
    }
    taxes = rows["Taxes (detail)"]
    assert taxes[2].value == 30_000 and taxes[5].value == "yes"
    utilities = rows["Utilities (detail)"]
    assert utilities[1].value == "per_unit"
    assert 'IF(B' in str(utilities[4].value)  # basis-resolving formula
    assert _input_cell(wb, "Total units") == 50
    # pct_of_egi lines fold into the fee cell, not the block.
    assert _input_cell(wb, "Mgmt fee % of EGI") == pytest.approx(0.03)


def test_development_workbook_has_formula_draws():
    from tests.parity.export_case import DEVELOPMENT_SCURVE_INPUTS

    content, warnings = build_model_workbook(DEVELOPMENT_SCURVE_INPUTS)
    wb = openpyxl.load_workbook(BytesIO(content))
    draws = wb["Draws"]
    # Weights are literal values; costs/draws/interest are formulas.
    assert isinstance(draws["B3"].value, float)
    assert str(draws["C3"].value).startswith("=(")
    assert str(draws["E3"].value).startswith("=C3-D3")
    assert "MIN(C3" in str(draws["D3"].value)
    assert str(draws["G3"].value).startswith("=IF(A3>=1,")
    assert any("S-curve" in w for w in warnings)
    # 12 construction months + month 0 row.
    assert draws.cell(row=14, column=8).value is not None
    assert draws.cell(row=15, column=1).value is None


def test_unit_mix_deals_collapse_income_with_a_warning(analytic):
    unit_mix_deal = {
        **analytic,
        "grossPotentialRent": 0,
        "unitMix": [{"unitType": "1BR", "unitCount": 50, "inPlaceRent": 1500}],
    }
    content, warnings = build_model_workbook(unit_mix_deal)
    assert any("collapsed to annual dollars" in w for w in warnings)
    wb = openpyxl.load_workbook(BytesIO(content))
    assert wb["Inputs"]["B6"].value == pytest.approx(50 * 1500 * 12)
    # ...and the warning also lands on the Notes sheet.
    notes = [c.value for c in wb["Notes"]["A"] if c.value]
    assert any("collapsed" in str(n) for n in notes)


def test_router_returns_workbook_or_422():
    client = TestClient(app)
    analytic = json.loads((FIXTURES / "analytic_acquisition.json").read_text())

    ok = client.post("/api/generate/model", json={"values": analytic})
    assert ok.status_code == 200
    assert ok.headers["content-type"].startswith("application/vnd.openxml")
    assert "X-Generation-Warnings" in ok.headers

    unsupported = client.post(
        "/api/generate/model",
        json={"values": {**analytic, "irrConvention": "xirr"}},
    )
    assert unsupported.status_code == 422
    assert "XIRR" in unsupported.json()["detail"]

    assert client.post("/api/generate/model", json={"values": {}}).status_code == 422


def test_output_cells_map_matches_the_sheet(analytic):
    wb = _workbook(analytic)
    out = wb["Outputs"]
    for output_id, cell in OUTPUT_CELLS.items():
        row = int(cell[1:])
        assert out.cell(row=row, column=6).value == output_id
        assert str(out[cell].value).startswith("=")


def test_export_module_importable_from_run():
    # run.py imports this lazily; keep the seam honest.
    from tests.parity.export_case import AMORTIZING_GROWTH_INPUTS, run_export_parity  # noqa: F401

    assert excel_model_export.unsupported_features(AMORTIZING_GROWTH_INPUTS) == []
