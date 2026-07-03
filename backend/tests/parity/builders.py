"""Synthetic parity templates: real xlsx workbooks whose formulas mirror the
native engine's math exactly for their input shapes, so any divergence between
the openpyxl+LibreOffice path and the engine is a genuine bug in one of them.

Both templates deliberately exercise the injection layer's audit fixes:
- a workbook-scoped named range and a SHEET-scoped named range (M1),
- a mapping that targets the non-anchor cell of a merged range (H5).

Shape constraints that make exact formula mirroring possible:
- acquisition: interest-only debt for the whole 60-month hold, flat growth
  -> every cash flow is closed-form and IRR ranges are constant vectors.
- development: constructionMonths = 0 (all costs at close), zero origination
  fee -> no capitalized interest, LTV governs the takeout sizing
  (dscr/debt-yield constraints set to 0 in the paired inputs).
"""

import openpyxl
from openpyxl.workbook.defined_name import DefinedName

HOLD_MONTHS = 60  # both templates hard-code a 5-year monthly grid


def build_acquisition_template(path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inputs"
    labels = [
        "Purchase Price", "Gross Potential Rent", "Vacancy %", "Opex (annual)",
        "Loan Amount", "Interest Rate", "Exit Cap Rate", "Hold (yrs)", "Cost of Sale %",
    ]
    for i, label in enumerate(labels, start=1):
        ws.cell(row=i, column=1, value=label)
    ws["B8"] = 5  # informational; the Calc grid is fixed at 60 months
    # Merged range: the mapping for costOfSalePct targets C9 (non-anchor).
    ws.merge_cells("B9:C9")

    calc = wb.create_sheet("Calc")
    calc["A1"] = "=Inputs!B2*(1-Inputs!B3)-Inputs!B4"  # annual NOI
    calc["A2"] = "=A1/12"  # monthly NOI
    calc["A3"] = "=Inputs!B5*Inputs!B6/12"  # monthly IO debt service
    calc["A4"] = "=A1/Inputs!B7"  # terminal value (forward NOI = NOI, flat)
    calc["A5"] = "=A4*(1-Inputs!B9)-Inputs!B5"  # net sale proceeds

    # Monthly vectors, rows 1..61 = t0..month 60.
    calc["C1"] = "=-(Inputs!B1-Inputs!B5)"  # equity out at close
    calc["D1"] = "=-Inputs!B1"
    for row in range(2, HOLD_MONTHS + 1):  # months 1..59
        calc.cell(row=row, column=3, value="=$A$2-$A$3")
        calc.cell(row=row, column=4, value="=$A$2")
    calc.cell(row=HOLD_MONTHS + 1, column=3, value="=$A$2-$A$3+$A$5")
    calc.cell(row=HOLD_MONTHS + 1, column=4, value="=$A$2+$A$4*(1-Inputs!B9)")

    calc["F1"] = "=(1+IRR(C1:C61))^12-1"  # levered IRR, engine annualization
    calc["F2"] = "=(1+IRR(D1:D61))^12-1"  # unlevered IRR
    calc["F3"] = '=SUMIF(C1:C61,">0")/-SUMIF(C1:C61,"<0")'  # equity multiple
    calc["F4"] = "=A4"
    calc["F5"] = "=A5"
    calc["F6"] = "=A1/Inputs!B1"  # going-in cap
    calc["F7"] = "=A1/Inputs!B1"  # yield on cost
    calc["F8"] = "=A2/A3"  # min DSCR
    calc["F9"] = "=A2/A3"  # avg DSCR
    calc["F10"] = "=A1/Inputs!B5"  # debt yield
    calc["F11"] = "=Inputs!B5/Inputs!B1"  # LTV
    calc["F12"] = "=Inputs!B5/Inputs!B1"  # LTC
    calc["F13"] = "=Inputs!B6"  # loan constant (pure IO)
    calc["F14"] = "=(Inputs!B4+Inputs!B5*Inputs!B6)/Inputs!B2"  # break-even ratio
    calc["F15"] = "=($A$2-$A$3)*12/(Inputs!B1-Inputs!B5)"  # year-1 CoC
    calc["F16"] = "=SUM(C1:C61)"  # total profit

    # Workbook-scoped name (standard) + sheet-scoped name (the M1 case).
    wb.defined_names["PurchasePrice"] = DefinedName(
        "PurchasePrice", attr_text="Inputs!$B$1"
    )
    ws.defined_names["LocalExitCap"] = DefinedName(
        "LocalExitCap", attr_text="Inputs!$B$7"
    )
    wb.save(path)


def build_development_template(path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inputs"
    labels = [
        "Land Cost", "Hard Costs", "Soft Costs", "Contingency %", "Developer Fee %",
        "Gross Potential Rent", "Vacancy %", "Opex (annual)", "LTC", "Interest Rate",
        "Exit Cap Rate", "Cost of Sale %",
    ]
    for i, label in enumerate(labels, start=1):
        ws.cell(row=i, column=1, value=label)
    # Merged range: vacancy mapping targets C7 (non-anchor).
    ws.merge_cells("B7:C7")

    calc = wb.create_sheet("Calc")
    calc["A1"] = "=(Inputs!B2+Inputs!B3)*Inputs!B4"  # contingency
    calc["A2"] = "=(Inputs!B2+Inputs!B3+A1)*Inputs!B5"  # developer fee
    calc["A3"] = "=Inputs!B1+Inputs!B2+Inputs!B3+A1+A2"  # TDC ex financing
    calc["A4"] = "=A3*(1-Inputs!B9)"  # equity (equity-first, fee = 0)
    calc["A5"] = "=A3*Inputs!B9"  # construction balance at close
    calc["A6"] = "=Inputs!B6*(1-Inputs!B7)-Inputs!B8"  # stabilized NOI
    calc["A7"] = "=A6/Inputs!B11"  # stabilized value
    calc["A8"] = "=A7*Inputs!B9"  # perm loan (LTV governs by construction)
    calc["A9"] = "=A8-A5"  # refi delta at takeout (month 1)
    calc["A10"] = "=A8*Inputs!B10/12"  # monthly IO perm debt service
    calc["A11"] = "=A6/Inputs!B11*(1-Inputs!B12)-A8"  # net sale proceeds

    calc["C1"] = "=-$A$4"  # equity at close
    calc["D1"] = "=-$A$3"  # unlevered: full budget at close
    calc["C2"] = "=$A$6/12-$A$10+$A$9"  # month 1: ops + refi delta
    calc["D2"] = "=$A$6/12"
    for row in range(3, HOLD_MONTHS + 1):
        calc.cell(row=row, column=3, value="=$A$6/12-$A$10")
        calc.cell(row=row, column=4, value="=$A$6/12")
    calc.cell(row=HOLD_MONTHS + 1, column=3, value="=$A$6/12-$A$10+$A$11")
    calc.cell(
        row=HOLD_MONTHS + 1, column=4,
        value="=$A$6/12+$A$6/Inputs!B11*(1-Inputs!B12)",
    )

    calc["F1"] = "=(1+IRR(C1:C61))^12-1"
    calc["F2"] = "=(1+IRR(D1:D61))^12-1"
    calc["F3"] = '=SUMIF(C1:C61,">0")/-SUMIF(C1:C61,"<0")'
    calc["F4"] = "=A6/Inputs!B11"  # terminal value
    calc["F5"] = "=A11"
    calc["F6"] = "=A6/A3"  # going-in cap = yield on cost for development
    calc["F7"] = "=A6/A3"
    calc["F8"] = "=A6/(A8*Inputs!B10)"  # min DSCR (IO)
    calc["F9"] = "=A6/(A8*Inputs!B10)"
    calc["F10"] = "=A6/A8"  # debt yield
    calc["F11"] = "=A8/A7"  # LTV (= the LTC input by construction)
    calc["F12"] = "=A8/A3"  # LTC output: perm / cost basis
    calc["F13"] = "=Inputs!B10"  # loan constant (pure IO)
    calc["F14"] = "=A6/A3-Inputs!B11"  # development spread (decimal)
    calc["F15"] = "=SUM(C1:C61)"  # total profit

    wb.defined_names["LandCost"] = DefinedName("LandCost", attr_text="Inputs!$B$1")
    ws.defined_names["LocalExitCap"] = DefinedName(
        "LocalExitCap", attr_text="Inputs!$B$11"
    )
    wb.save(path)
