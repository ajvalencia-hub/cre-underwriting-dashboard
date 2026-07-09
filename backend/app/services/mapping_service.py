import json
import re
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from app.config import INPUT_SCHEMA_PATH

MIN_KEYWORD_LEN = 4
MAX_LABEL_SEARCH_ROWS = 400
MAX_LABEL_SEARCH_COLS = 50


def _normalize(name: str) -> str:
    return re.sub(r"[^a-z0-9]", "", name.lower())


def load_flat_fields(include_outputs: bool = False) -> list[dict]:
    with open(INPUT_SCHEMA_PATH, encoding="utf-8") as f:
        schema = json.load(f)
    fields = []
    for section in schema["sections"]:
        fields.extend(section["fields"])
    if include_outputs:
        fields.extend(schema.get("outputs", []))
    return fields


def load_output_fields() -> list[dict]:
    with open(INPUT_SCHEMA_PATH, encoding="utf-8") as f:
        schema = json.load(f)
    return schema.get("outputs", [])


def _match_by_named_range(fields: list[dict], named_ranges: list[dict]) -> dict:
    nr_by_norm: dict[str, dict] = {}
    for nr in named_ranges:
        nr_by_norm.setdefault(_normalize(nr["name"]), nr)

    mappings: dict[str, dict] = {}
    for field in fields:
        nr = nr_by_norm.get(_normalize(field["id"]))
        if nr is None:
            continue

        if field["type"] == "table":
            mappings[field["id"]] = {
                "target": "table",
                "anchor": nr["ref"],
                "sheet": nr["sheet"],
                "columnOrder": [c["id"] for c in field.get("columns", [])],
                "source": "auto",
            }
        elif field["type"] == "keyvalue":
            continue
        else:
            mappings[field["id"]] = {
                "target": "namedRange",
                "ref": nr["name"],
                "sheet": None,
                "source": "auto",
            }

    return mappings


def _find_adjacent_value_cell(ws, label_row: int, label_col: int, max_col: int):
    """Given a cell that looks like a field label, guess where its value lives.

    Checks up to 3 cells to the right and returns the first one that is either
    empty (an unfilled input slot) or holds a non-string value (a number/date/
    formula) — i.e. the first cell that doesn't look like another label. This
    matches the common 'Label | Value' or 'Label | unit | Value' layout used in
    most underwriting assumption sheets.
    """
    for offset in (1, 2, 3):
        col = label_col + offset
        if col > max_col:
            break
        value = ws.cell(row=label_row, column=col).value
        if value is None or not isinstance(value, str):
            return label_row, col
    if label_col + 1 <= max_col:
        return label_row, label_col + 1
    return label_row + 1, label_col


def _keywords_for(field: dict) -> tuple[str, str]:
    return _normalize(field.get("label", "")), _normalize(field["id"])


def _build_mapping_entry(field: dict, sheet_title: str, coord: str) -> dict | None:
    if field["type"] == "table":
        return {
            "target": "table",
            "anchor": coord,
            "sheet": sheet_title,
            "columnOrder": [c["id"] for c in field.get("columns", [])],
            "source": "auto",
        }
    if field["type"] == "keyvalue":
        return None
    return {
        "target": "cell",
        "ref": f"{sheet_title}!{coord}",
        "sheet": sheet_title,
        "source": "auto",
    }


def _match_by_label_search(template_path: Path, fields: list[dict]) -> dict:
    """Scan cell text across the workbook for labels that name a field
    ('Purchase Price', 'Interest Rate', ...) and map the field to whatever
    cell sits next to that label — the universal fallback for templates that
    don't define named ranges.
    """
    if not fields:
        return {}

    wb = openpyxl.load_workbook(template_path, data_only=False)

    candidates: list[tuple[object, int, int, str]] = []
    for ws in wb.worksheets:
        max_row = min(ws.max_row or 0, MAX_LABEL_SEARCH_ROWS)
        max_col = min(ws.max_column or 0, MAX_LABEL_SEARCH_COLS)
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                value = ws.cell(row=r, column=c).value
                if not isinstance(value, str) or value.startswith("="):
                    continue
                norm = _normalize(value)
                if len(norm) < 3:
                    continue
                candidates.append((ws, r, c, norm))

    remaining = {f["id"]: f for f in fields}
    mappings: dict[str, dict] = {}

    def sweep(is_match) -> None:
        for ws, r, c, norm in candidates:
            if not remaining:
                return
            for field_id, field in list(remaining.items()):
                label_kw, id_kw = _keywords_for(field)
                if not is_match(norm, label_kw, id_kw):
                    continue
                max_col = min(ws.max_column or c, MAX_LABEL_SEARCH_COLS)
                target_row, target_col = _find_adjacent_value_cell(ws, r, c, max_col)
                coord = f"{get_column_letter(target_col)}{target_row}"
                entry = _build_mapping_entry(field, ws.title, coord)
                if entry is not None:
                    mappings[field_id] = entry
                del remaining[field_id]
                break

    # Pass 1: exact match (highest confidence) — cell text normalizes to
    # exactly the field's label or id.
    sweep(lambda norm, label_kw, id_kw: norm == label_kw or norm == id_kw)

    # Pass 2: substring match, only for keywords long enough to be meaningful,
    # to catch label wording that doesn't exactly match the schema label
    # (e.g. a cell reading "Exit Cap Rate" against field id "exitCapRatePct").
    if remaining:
        sweep(
            lambda norm, label_kw, id_kw: (
                (len(label_kw) >= MIN_KEYWORD_LEN and (label_kw in norm or norm in label_kw))
                or (len(id_kw) >= MIN_KEYWORD_LEN and (id_kw in norm or norm in id_kw))
            )
        )

    wb.close()
    return mappings


def auto_match(named_ranges: list[dict], template_path: Path | None = None) -> dict:
    """Match input-schema fields (and computed-output fields) to cells in the template.

    Two passes, named ranges first since they're explicit and stable across
    row/column edits:
      1. Named-range match by normalized name.
      2. Keyword/label search across cell text for anything left unmatched —
         finds a label like 'Purchase Price' and maps the field to the cell
         next to it. This is what makes auto-match useful even for templates
         with no named ranges at all.
    """
    fields = load_flat_fields(include_outputs=True)

    mappings = _match_by_named_range(fields, named_ranges)

    if template_path is not None:
        remaining_fields = [f for f in fields if f["id"] not in mappings]
        mappings.update(_match_by_label_search(template_path, remaining_fields))

    return mappings


def compute_unmapped_required(mappings: dict) -> list[str]:
    fields = load_flat_fields()
    return [f["id"] for f in fields if f.get("required") and f["id"] not in mappings]
