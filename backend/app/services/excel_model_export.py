"""Native Excel model export (H11, widened by I7): a formula-live workbook
whose cells mirror the engine's math, so the downloaded file recalculates on
its own.

Supported shapes — acquisitions AND developments with GPR-based income,
flat or line-item expenses (incl. non-ad-valorem), simple debt (IO then
level amortization; construction-to-perm for developments), single-cap
exit, pro-rata equity, periodic-monthly IRR. Anything whose math can't be
mirrored formula-for-formula raises UnsupportedModelFeatures listing the
blockers — a silently wrong model must never escape (see DECISIONS.md).

Deliberate value-not-formula cells, each flagged in the Notes sheet:
- Loan amount (acquisition) / permanent takeout amount (development): the
  engine's constraint-based sizing written as the sized VALUE.
- Annual GPR / other income: unit-mix or per-SF sections collapse to the
  same annual dollars the engine uses.
- Development S-curve weights: literal values on the Draws sheet (the
  cosine ogive isn't worth mirroring as a formula); draws, equity-first
  split, fee, and capitalized interest are FORMULAS over them.
- Timeline constants (construction months, takeout month, ramp) from the
  engine's timeline.

Sheets: Inputs (scalars + the Expenses block), Draws (development only),
Model (monthly grid), Debt (schedule tied to Model by reference), Outputs
(flows + metrics), Notes.
"""

from io import BytesIO

import openpyxl

from app.services.proforma import development, engine, leases, operations
from app.services.proforma.operations import (
    EXPENSE_DOLLAR_FIELDS,
    RECOVERABLE_EXPENSE_FIELDS,
)
from app.services.proforma.timeline import build_timeline


class UnsupportedModelFeatures(Exception):
    def __init__(self, features: list[str]):
        self.features = features
        super().__init__(
            "The Excel model export can't mirror these features as formulas: "
            + "; ".join(features)
        )


# Outputs sheet cells the parity harness reads, keyed by schema output id.
OUTPUT_CELLS = {
    "unleveredIrr": "G2",
    "leveredIrr": "G3",
    "equityMultiple": "G4",
    "cashOnCashYear1": "G5",
    "goingInCapRate": "G6",
    "yieldOnCost": "G7",
    "terminalValue": "G8",
    "netSaleProceeds": "G9",
    "totalProfit": "G10",
    "npv": "G11",
    "minDscr": "G12",
    "debtYield": "G13",
    "ltv": "G14",
    "loanConstant": "G15",
}

_EXPENSE_LABELS = {
    "realEstateTaxes": "Real estate taxes",
    "insurance": "Insurance",
    "utilities": "Utilities",
    "repairsMaintenance": "Repairs & maintenance",
    "payroll": "Payroll",
    "generalAdmin": "General & admin",
    "replacementReserves": "Replacement reserves",
}
_DETAIL_LABELS = {
    "taxes": "Taxes (detail)",
    "insurance": "Insurance (detail)",
    "utilities": "Utilities (detail)",
    "repairs_maintenance": "Repairs & maintenance (detail)",
    "payroll": "Payroll (detail)",
    "ga": "General & admin (detail)",
    "other": "Other opex (detail)",
}


def _num(inputs: dict, key: str, default: float = 0.0) -> float:
    value = inputs.get(key)
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return default
    return float(value)


def unsupported_features(inputs: dict) -> list[str]:
    features = []
    if leases.has_leases(inputs):
        features.append("commercial lease-level rent rolls (escalations/recoveries/rollover)")
    if inputs.get("waterfallTiers"):
        features.append("promote waterfall tiers")
    if inputs.get("irrConvention") == "xirr":
        features.append("XIRR date-based IRR convention")
    if inputs.get("useReassessedTaxes"):
        features.append("reassessed property taxes (separate tax growth clock)")
    if any(
        isinstance(r, dict) and _num(r, "unitsToReno") > 0
        for r in (inputs.get("renovationProgram") or [])
    ):
        features.append("value-add renovation program (per-unit-type downtime/premium schedule)")
    if _num(inputs, "assetMgmtFeePct") > 0:
        features.append("asset management fee (new partnership-level fee timing)")
    if inputs.get("juniorTrancheKind") in ("mezz", "pref_equity") and (
        _num(inputs, "juniorTrancheAmount") > 0
        or _num(inputs, "juniorTrancheTotalLtcPct") > 0
    ):
        features.append("junior tranche (mezzanine debt / preferred equity)")
    if (inputs.get("dealType") or "acquisition") == "development":
        hold_years = _num(inputs, "holdPeriodYears", 5)
        timeline, _ = build_timeline(
            "development", hold_years,
            construction_months=_num(inputs, "constructionMonths") or None,
            lease_up_months=_num(inputs, "leaseUpMonths") or None,
            stabilization_month=_num(inputs, "stabilizationMonth") or None,
        )
        if timeline.stabilization_month > timeline.total_months:
            features.append(
                "development sold before stabilization (no permanent takeout occurs)"
            )
    return features


def _expense_rows(inputs: dict) -> tuple[list[dict], float, list[str]]:
    """The Expenses block: rows of {label, basis, amount, growth,
    recoverable} where `amount` resolves per basis in-sheet. Returns
    (rows, egi_pct_total, warnings)."""
    warnings: list[str] = []
    expense_growth = (
        _num(inputs, "expenseGrowthPct") if inputs.get("expenseGrowthMode") != "flat" else 0.0
    )

    rows: list[dict] = []
    egi_pct_total = 0.0
    if operations.has_opex_detail(inputs):
        for raw in inputs.get("opexLineItems") or []:
            if not (isinstance(raw, dict) and _num(raw, "amount") > 0):
                continue
            basis = raw.get("basis") or "annual_total"
            if basis == "pct_of_egi":
                egi_pct_total += _num(raw, "amount")
                continue
            category = raw.get("category") or "other"
            growth = (
                _num(raw, "growthPct") if raw.get("growthPct") is not None else expense_growth
            )
            rows.append({
                "label": _DETAIL_LABELS.get(category, category),
                "basis": basis,
                "amount": _num(raw, "amount"),
                "growth": growth,
                "recoverable": raw.get("recoverable") in (True, "yes", "true", 1),
            })
    else:
        egi_pct_total = _num(inputs, "managementFeePct")
        for field in EXPENSE_DOLLAR_FIELDS:
            rows.append({
                "label": _EXPENSE_LABELS.get(field, field),
                "basis": "annual_total",
                "amount": _num(inputs, field),
                "growth": expense_growth,
                "recoverable": field in RECOVERABLE_EXPENSE_FIELDS,
            })

    # I5: non-ad-valorem assessments — its own growth clock, in both modes.
    nav = _num(inputs, "nonAdValoremTaxes")
    if nav > 0:
        nav_flag = inputs.get("nonAdValoremRecoverable")
        rows.append({
            "label": "Non-ad-valorem assessments",
            "basis": "annual_total",
            "amount": nav,
            "growth": (
                _num(inputs, "nonAdValoremGrowthPct")
                if inputs.get("nonAdValoremGrowthPct") is not None
                else expense_growth
            ),
            "recoverable": True if nav_flag is None else bool(nav_flag),
        })
    return rows, egi_pct_total, warnings


def build_model_workbook(inputs: dict) -> tuple[bytes, list[str]]:
    """Returns (xlsx bytes, warnings). Raises UnsupportedModelFeatures or
    engine.InsufficientInputsError."""
    blockers = unsupported_features(inputs)
    if blockers:
        raise UnsupportedModelFeatures(blockers)

    is_dev = (inputs.get("dealType") or "acquisition") == "development"

    # One engine compute for the sized loan (and to fail early on bad inputs).
    result = engine.compute(inputs)
    debt_block = result.get("debt") or {}
    loan_amount = float(debt_block.get("loanAmount") or 0.0)

    hold_years = _num(inputs, "holdPeriodYears", 5)
    timeline, _tl_warnings = build_timeline(
        inputs.get("dealType") or "acquisition", hold_years,
        construction_months=_num(inputs, "constructionMonths") or None,
        lease_up_months=_num(inputs, "leaseUpMonths") or None,
        stabilization_month=_num(inputs, "stabilizationMonth") or None,
    )
    hold_months = timeline.total_months
    cm = timeline.construction_months
    takeout = min(timeline.stabilization_month, hold_months + 1) if is_dev else 1
    ramp = max(1, timeline.stabilization_month - cm - 1)
    total_rows = hold_months + 12  # forward window for the exit cap

    annual_gpr, annual_other, gpr_source, _ = operations.annual_gpr_and_other_income(inputs)
    warnings: list[str] = []
    if gpr_source != "grossPotentialRent":
        warnings.append(
            f"Income comes from a {gpr_source} section — collapsed to annual dollars "
            "(the rows themselves are not formula-modeled)."
        )
    if loan_amount > 0:
        warnings.append(
            (
                "Permanent takeout ${:,.0f} is the app's constraint-based sizing "
                "(governed by {}) written as a value."
                if is_dev
                else "Loan amount ${:,.0f} is the app's constraint-based sizing "
                "(governed by {}) written as a value."
            ).format(loan_amount, debt_block.get("governingConstraint", "ltv"))
        )
    if is_dev:
        warnings.append(
            "Construction S-curve weights are literal values on the Draws sheet; "
            "draws, equity-first split, fee, and capitalized interest are formulas "
            "over them."
        )

    rent_growth = (
        _num(inputs, "rentGrowthPct") if inputs.get("rentGrowthMode") != "flat" else 0.0
    )
    expense_rows, egi_pct_total, exp_warnings = _expense_rows(inputs)
    warnings.extend(exp_warnings)

    total_units = sum(
        _num(r, "unitCount") for r in (inputs.get("unitMix") or []) if isinstance(r, dict)
    )
    total_sf = _num(inputs, "rentableSf") or _num(inputs, "officeRentableSf")

    wb = openpyxl.Workbook()
    R: dict[str, str] = {}  # semantic name -> absolute cell ref

    # ---- Inputs sheet -----------------------------------------------------
    ws = wb.active
    ws.title = "Inputs"
    row = 0

    def put(key: str, label: str, value):
        nonlocal row
        row += 1
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        R[key] = f"Inputs!$B${row}"

    if is_dev:
        put("land", "Land cost", _num(inputs, "landCost"))
        put("hard", "Hard costs", _num(inputs, "hardCosts"))
        put("soft", "Soft costs", _num(inputs, "softCosts"))
        put("contPct", "Contingency % (of hard+soft)", _num(inputs, "contingencyPct", 0.05))
        put("feePct", "Developer fee % (of hard+soft+contingency)", _num(inputs, "developerFeePct", 0.04))
        put("contingency", "Contingency", f"=({R['hard']}+{R['soft']})*{R['contPct']}")
        put("devFee", "Developer fee", f"=({R['hard']}+{R['soft']}+{R['contingency']})*{R['feePct']}")
        put("totalExFin", "Total budget (ex financing)",
            f"={R['land']}+{R['hard']}+{R['soft']}+{R['contingency']}+{R['devFee']}")
        put("ltc", "LTC", _num(inputs, "ltvOrLtc", 0.65))
        put("equityTarget", "Equity (funds first)", f"={R['totalExFin']}*(1-{R['ltc']})")
        put("spread", "Refi/perm rate spread", _num(inputs, "refiRateSpreadPct"))
        put("refiCostsPct", "Refi costs % (of perm loan)", _num(inputs, "refiCostsPct"))
    else:
        put("price", "Purchase price", _num(inputs, "purchasePrice"))
        put("closingPct", "Closing costs %", _num(inputs, "closingCostsPct"))
        put("acqFeePct", "Acquisition fee %", _num(inputs, "acquisitionFeePct"))
        put("dd", "Due diligence", _num(inputs, "dueDiligenceCosts"))
        put("capex", "Day-1 capex", _num(inputs, "dayOneCapex"))

    put("gpr", "Annual GPR", annual_gpr)
    put("other", "Annual other income", annual_other)
    put("vac", "Vacancy %", _num(inputs, "vacancyPct", 0.05))
    put("stabOcc", "Stabilized occupancy", f"=1-{R['vac']}")
    put("credit", "Credit loss %", _num(inputs, "creditLossPct"))
    put("rentG", "Rent growth %/yr", rent_growth)
    put("mgmtPct", "Mgmt fee % of EGI", egi_pct_total)
    put("loan", "Perm loan (sized by app)" if is_dev else "Loan amount (sized by app)", loan_amount)
    put("rate", "Interest rate", _num(inputs, "interestRate", 0.065))
    if is_dev:
        put("permRate", "Perm rate (rate + spread)", f"={R['rate']}+{R['spread']}")
    else:
        R["permRate"] = R["rate"]
    put("amort", "Amortization (yrs)", _num(inputs, "amortYears", 30))
    put("io", "IO months", int(_num(inputs, "ioMonths")))
    put("origFee", "Origination fee %", _num(inputs, "originationFeePct"))
    put("hold", "Hold (months)", hold_months)
    put("cm", "Construction months", cm)
    put("takeout", "Perm takeout month", takeout)
    put("ramp", "Lease-up ramp (months)", ramp)
    put("exitCap", "Exit cap rate", _num(inputs, "exitCapRatePct"))
    put("cos", "Cost of sale %", _num(inputs, "costOfSalePct"))
    put("disc", "Discount rate", _num(inputs, "discountRatePct", 0.10))
    put("inPlaceNoi", "In-place NOI (0 = use year 1)", _num(inputs, "inPlaceNoi"))
    put("totalUnits", "Total units (per_unit basis)", total_units)
    put("totalSf", "Total SF (psf basis)", total_sf)

    # Expenses block: label | basis | amount | growth | resolved annual | note
    row += 2
    ws.cell(row=row, column=1, value="OPERATING EXPENSES")
    ws.cell(row=row, column=2, value="basis")
    ws.cell(row=row, column=3, value="amount")
    ws.cell(row=row, column=4, value="growth %/yr")
    ws.cell(row=row, column=5, value="annual (resolved)")
    ws.cell(row=row, column=6, value="recoverable")
    exp_first = row + 1
    for item in expense_rows:
        row += 1
        ws.cell(row=row, column=1, value=item["label"])
        ws.cell(row=row, column=2, value=item["basis"])
        ws.cell(row=row, column=3, value=item["amount"])
        ws.cell(row=row, column=4, value=item["growth"])
        ws.cell(
            row=row, column=5,
            value=(
                f'=C{row}*IF(B{row}="per_unit",IF({R["totalUnits"]}>0,{R["totalUnits"]},1),'
                f'IF(B{row}="psf",IF({R["totalSf"]}>0,{R["totalSf"]},1),1))'
            ),
        )
        ws.cell(row=row, column=6, value="yes" if item["recoverable"] else "no")
    exp_last = row if expense_rows else exp_first
    if not expense_rows:  # keep the SUMPRODUCT ranges valid
        row += 1
        ws.cell(row=row, column=1, value="(none)")
        ws.cell(row=row, column=3, value=0)
        ws.cell(row=row, column=4, value=0)
        ws.cell(row=row, column=5, value=0)
        exp_last = row
    exp_annual = f"Inputs!$E${exp_first}:$E${exp_last}"
    exp_growth = f"Inputs!$D${exp_first}:$D${exp_last}"

    # ---- Draws sheet (development) -----------------------------------------
    if is_dev:
        draws = wb.create_sheet("Draws")
        for col, header in enumerate(
            ["Month", "S-curve weight", "Cost", "Equity funded", "Loan draw",
             "Fee (capitalized)", "Interest (capitalized)", "End balance"],
            start=1,
        ):
            draws.cell(row=1, column=col, value=header)
        s_weights = development.s_curve_weights(cm)
        line = f"1/{cm}" if cm > 0 else "0"
        for m in range(0, cm + 1):
            r = m + 2
            draws.cell(row=r, column=1, value=m)
            if m == 0:
                draws.cell(row=r, column=2, value=0.0)
                draws.cell(row=r, column=3, value=f"={R['land']}")
            else:
                draws.cell(row=r, column=2, value=s_weights[m - 1])  # literal
                draws.cell(
                    row=r, column=3,
                    value=f"=({R['hard']}+{R['contingency']})*B{r}"
                          f"+({R['soft']}+{R['devFee']})*{line}",
                )
            prior_equity = f"SUM(D$2:D{r - 1})" if r > 2 else "0"
            draws.cell(
                row=r, column=4,
                value=f"=MAX(0,MIN(C{r},{R['equityTarget']}-{prior_equity}))",
            )
            draws.cell(row=r, column=5, value=f"=C{r}-D{r}")
            prior_draws = f"SUM(E$2:E{r - 1})" if r > 2 else "0"
            draws.cell(
                row=r, column=6,
                value=f"=IF(AND(E{r}>0,{prior_draws}=0),{R['origFee']}*E{r},0)",
            )
            prev_bal = f"H{r - 1}" if r > 2 else "0"
            pre_interest = f"({prev_bal}+F{r}+E{r})"
            draws.cell(
                row=r, column=7,
                value=f"=IF(A{r}>=1,{pre_interest}*{R['rate']}/12,0)",
            )
            draws.cell(row=r, column=8, value=f"={pre_interest}+G{r}")
        draws_last = cm + 2
        R["drawsEnd"] = f"Draws!$H${draws_last}"
        R["capInterest"] = f"SUM(Draws!$G$2:$G${draws_last})"
        R["capFee"] = f"SUM(Draws!$F$2:$F${draws_last})"

    # ---- Model sheet: rows 2..total_rows+1 = months 1..total_rows ----------
    model = wb.create_sheet("Model")
    headers = [
        "Month", "Op month", "Occupancy", "GPR", "Vacancy loss", "Credit loss",
        "Other income", "EGI", "Fixed opex", "Mgmt fee", "NOI",
        "Begin balance", "Interest", "Principal", "End balance", "Debt service",
        "DSCR", "Unlevered CF", "Levered CF",
    ]
    for col, header in enumerate(headers, start=1):
        model.cell(row=1, column=col, value=header)

    for m in range(1, total_rows + 1):
        r = m + 1
        g = f"INT(($B{r}-1)/12)"
        model.cell(row=r, column=1, value=m)
        model.cell(row=r, column=2, value=f"=A{r}-{R['cm']}")
        model.cell(
            row=r, column=3,
            value=f"=IF($B{r}<1,0,{R['stabOcc']}*MIN(1,$B{r}/{R['ramp']}))",
        )
        model.cell(row=r, column=4, value=f"=IF($B{r}<1,0,{R['gpr']}/12*(1+{R['rentG']})^{g})")
        model.cell(row=r, column=5, value=f"=D{r}*(1-C{r})")
        model.cell(row=r, column=6, value=f"=D{r}*C{r}*{R['credit']}")
        model.cell(
            row=r, column=7,
            value=f"=IF($B{r}<1,0,{R['other']}/12*(1+{R['rentG']})^{g}"
                  f"*IF({R['stabOcc']}>0,C{r}/{R['stabOcc']},0))",
        )
        model.cell(row=r, column=8, value=f"=D{r}-E{r}-F{r}+G{r}")
        model.cell(
            row=r, column=9,
            value=f"=IF($B{r}<1,0,SUMPRODUCT({exp_annual}/12,(1+{exp_growth})^{g}))",
        )
        model.cell(row=r, column=10, value=f"=H{r}*{R['mgmtPct']}")
        model.cell(row=r, column=11, value=f"=H{r}-I{r}-J{r}")

        if m > hold_months:
            continue  # forward window: income only, no debt or cash flows

        exit_u = f"+IF($A{r}={R['hold']},Outputs!$B$6,0)"
        exit_l = f"+IF($A{r}={R['hold']},Outputs!$B$6-O{r},0)"

        if not is_dev:
            begin = R["loan"] if m == 1 else f"O{r - 1}"
            model.cell(row=r, column=12, value=f"={begin}")
            model.cell(row=r, column=13, value=f"=L{r}*{R['rate']}/12")
            model.cell(
                row=r, column=14,
                value=f"=IF($A{r}<={R['io']},0,MAX(0,MIN(Outputs!$B$1-M{r},L{r})))",
            )
            model.cell(row=r, column=15, value=f"=L{r}-N{r}")
            model.cell(row=r, column=16, value=f"=M{r}+N{r}")
            model.cell(row=r, column=17, value=f'=IF(P{r}>0,K{r}/P{r},"")')
            model.cell(row=r, column=18, value=f"=K{r}{exit_u}")
            model.cell(row=r, column=19, value=f"=K{r}-P{r}{exit_l}")
            continue

        # Development phases.
        if m <= cm:
            draws_row = m + 2
            model.cell(row=r, column=15, value=f"=Draws!$H${draws_row}")
            model.cell(row=r, column=18, value=f"=K{r}-Draws!$C${draws_row}{exit_u}")
            model.cell(row=r, column=19, value=f"=-Draws!$D${draws_row}")
        elif m < takeout:
            begin = R["drawsEnd"] if m == cm + 1 else f"O{r - 1}"
            model.cell(row=r, column=12, value=f"={begin}")
            model.cell(row=r, column=13, value=f"=L{r}*{R['rate']}/12")
            model.cell(row=r, column=15, value=f"=MAX(0,L{r}+M{r}-K{r})")
            model.cell(row=r, column=18, value=f"=K{r}{exit_u}")
            model.cell(row=r, column=19, value="=0")
        else:
            if m == takeout:
                carry_end = (
                    R["drawsEnd"] if takeout == cm + 1 else f"O{r - 1}"
                )
                model.cell(row=r, column=12, value=f"={R['loan']}")
                refi = (
                    f"({R['loan']}-{carry_end}-{R['loan']}*{R['refiCostsPct']})"
                )
                model.cell(
                    row=r, column=19,
                    value=f"={refi}+K{r}-P{r}{exit_l}",
                )
            else:
                model.cell(row=r, column=12, value=f"=O{r - 1}")
                model.cell(row=r, column=19, value=f"=K{r}-P{r}{exit_l}")
            model.cell(row=r, column=13, value=f"=L{r}*{R['permRate']}/12")
            perm_month = f"($A{r}-{R['takeout']}+1)"
            model.cell(
                row=r, column=14,
                value=f"=IF({perm_month}<={R['io']},0,MAX(0,MIN(Outputs!$B$1-M{r},L{r})))",
            )
            model.cell(row=r, column=15, value=f"=L{r}-N{r}")
            model.cell(row=r, column=16, value=f"=M{r}+N{r}")
            model.cell(row=r, column=17, value=f'=IF(P{r}>0,K{r}/P{r},"")')
            model.cell(row=r, column=18, value=f"=K{r}{exit_u}")

    hold_row = hold_months + 1
    fwd_first, fwd_last = hold_row + 1, hold_row + 12

    # ---- Debt sheet: the schedule, tied to Model by reference --------------
    debt_ws = wb.create_sheet("Debt")
    for col, header in enumerate(
        ["Month", "Opening balance", "Interest", "Principal", "Debt service",
         "Closing balance", "DSCR"],
        start=1,
    ):
        debt_ws.cell(row=1, column=col, value=header)
    for m in range(1, hold_months + 1):
        r = m + 1
        debt_ws.cell(row=r, column=1, value=m)
        for col, model_col in ((2, "L"), (3, "M"), (4, "N"), (5, "P"), (6, "O"), (7, "Q")):
            debt_ws.cell(row=r, column=col, value=f"=Model!${model_col}${r}")

    # ---- Outputs sheet ------------------------------------------------------
    out = wb.create_sheet("Outputs")
    if is_dev:
        basis_formula = f"={R['totalExFin']}+{R['capInterest']}+{R['capFee']}"
        fees_formula = f"={R['capFee']}"
        equity_formula = f"={R['equityTarget']}"
        pmt_loan, pmt_rate = R["loan"], R["permRate"]
    else:
        basis_formula = (
            f"={R['price']}*(1+{R['closingPct']}+{R['acqFeePct']})+{R['dd']}+{R['capex']}"
        )
        fees_formula = f"={R['loan']}*{R['origFee']}"
        equity_formula = f"=B2-{R['loan']}+B3"
        pmt_loan, pmt_rate = R["loan"], R["rate"]

    helper_rows = [
        ("Monthly payment (PMT)",
         f"=IF({pmt_loan}<=0,0,IF({pmt_rate}=0,"
         f"{pmt_loan}/ROUND({R['amort']}*12,0),"
         f"{pmt_loan}*{pmt_rate}/12/(1-(1+{pmt_rate}/12)^-ROUND({R['amort']}*12,0))))"),
        ("Total cost basis", basis_formula),
        ("Loan fees", fees_formula),
        ("Initial equity", equity_formula),
        ("Terminal value (fwd-12 NOI / cap)",
         f"=SUM(Model!$K${fwd_first}:$K${fwd_last})/{R['exitCap']}"),
        ("Gross sale net of costs", f"=B5*(1-{R['cos']})"),
        ("Exit loan balance", f"=Model!$O${hold_row}"),
        ("Net sale proceeds", "=B6-B7"),
        ("Stabilized EGI (today's rents)",
         f"={R['gpr']}*(1-{R['vac']})*(1-{R['credit']})+{R['other']}"),
        ("Stabilized NOI",
         f"=B9-SUM({exp_annual})-B9*{R['mgmtPct']}"),
    ]
    for i, (label, formula) in enumerate(helper_rows, start=1):
        out.cell(row=i, column=1, value=label)
        out.cell(row=i, column=2, value=formula)

    # Cash-flow vectors: D = levered, E = unlevered; row 1 = close (t0).
    if is_dev:
        out["D1"] = "=-Draws!$D$2"
        out["E1"] = "=-Draws!$C$2"
    else:
        out["D1"] = "=-B4"
        out["E1"] = "=-B2"
    for m in range(1, hold_months + 1):
        out.cell(row=m + 1, column=4, value=f"=Model!S{m + 1}")
        out.cell(row=m + 1, column=5, value=f"=Model!R{m + 1}")
    flow_last = hold_months + 1
    lev = f"$D$1:$D${flow_last}"
    unlev = f"$E$1:$E${flow_last}"

    coc_exit_adj = "-B8" if hold_months == 12 else ""
    # Yield on cost divides by the engine's total_cost_basis: the development
    # basis already capitalizes the fee (B2); acquisitions add cash loan fees
    # on top (B2+B3).
    yoc = "=B10/B2" if is_dev else "=B10/(B2+B3)"
    if is_dev:
        going_in = "=B10/B2"
        ltv_formula = f"={R['loan']}/(B10/{R['exitCap']})"
        constant_io = f"{R['io']}>={R['hold']}-{R['takeout']}+1"
        constant_rate = R["permRate"]
    else:
        going_in = (
            f"=IF({R['inPlaceNoi']}>0,{R['inPlaceNoi']},SUM(Model!$K$2:$K$13))/{R['price']}"
        )
        ltv_formula = f"={R['loan']}/{R['price']}"
        constant_io = f"{R['io']}>={R['hold']}"
        constant_rate = R["rate"]

    metrics = {
        "unleveredIrr": f"=(1+IRR({unlev},0.005))^12-1",
        "leveredIrr": f"=(1+IRR({lev},0.005))^12-1",
        "equityMultiple": f'=SUMIF({lev},">0")/-SUMIF({lev},"<0")',
        "cashOnCashYear1": f'=(SUM(Model!$S$2:$S$13){coc_exit_adj})/-SUMIF({lev},"<0")',
        "goingInCapRate": going_in,
        "yieldOnCost": yoc,
        "terminalValue": "=B5",
        "netSaleProceeds": "=B8",
        "totalProfit": f"=SUM({lev})",
        "npv": f"=D1+NPV((1+{R['disc']})^(1/12)-1,$D$2:$D${flow_last})",
        "minDscr": f"=MIN(Model!$Q$2:$Q${hold_row})",
        "debtYield": f"=B10/{R['loan']}",
        "ltv": ltv_formula,
        "loanConstant": f"=IF({constant_io},{constant_rate},12*B1/{R['loan']})",
    }
    for i, (output_id, formula) in enumerate(metrics.items(), start=2):
        out.cell(row=i, column=6, value=output_id)
        assert OUTPUT_CELLS[output_id] == f"G{i}"
        out.cell(row=i, column=7, value=formula)

    # ---- Notes sheet ------------------------------------------------------
    notes = wb.create_sheet("Notes")
    notes["A1"] = "Exported by CRE Underwriting Dashboard — formula-live model"
    note_lines = warnings + [
        "Conventions: monthly periods, growth steps annually (anniversary) on "
        "the OPERATING clock, credit loss applies to occupied revenue, terminal "
        "value caps the FORWARD 12-month NOI, IRR is monthly annualized as "
        "(1+i)^12-1. Recoverable flags on the Expenses block are annotations "
        "(recoveries need lease-level modeling, which this export refuses).",
    ]
    for i, line in enumerate(note_lines, start=3):
        notes.cell(row=i, column=1, value=line)

    wb.calculation.fullCalcOnLoad = True
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), warnings
