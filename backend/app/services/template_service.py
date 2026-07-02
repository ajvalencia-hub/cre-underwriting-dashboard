import hashlib
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import IllegalCharacterError


def compute_file_hash(file_bytes: bytes) -> str:
    return hashlib.sha256(file_bytes).hexdigest()


def _safe_value(value: Any) -> Any:
    if value is None or isinstance(value, (int, float, str, bool)):
        return value
    return str(value)


def parse_workbook(path: Path) -> dict:
    """Read sheet dimensions and defined names (named ranges) from a workbook."""
    wb = openpyxl.load_workbook(path, data_only=False)

    sheets = [
        {"name": ws.title, "maxRow": ws.max_row, "maxCol": ws.max_column}
        for ws in wb.worksheets
    ]

    named_ranges = []
    for name, defined_name in wb.defined_names.items():
        if defined_name.type != "RANGE":
            continue
        try:
            for sheet_title, coord in defined_name.destinations:
                named_ranges.append(
                    {"name": name, "sheet": sheet_title, "ref": coord.replace("$", "")}
                )
        except (TypeError, ValueError):
            continue

    wb.close()
    return {"sheets": sheets, "namedRanges": named_ranges}


def get_sheet_grid(path: Path, sheet_name: str, max_rows: int = 60, max_cols: int = 30) -> dict:
    wb = openpyxl.load_workbook(path, data_only=False)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise KeyError(sheet_name)

    ws = wb[sheet_name]
    n_rows = min(ws.max_row, max_rows) if ws.max_row else 0
    n_cols = min(ws.max_column, max_cols) if ws.max_column else 0

    rows_out = []
    for r in range(1, n_rows + 1):
        row_cells = []
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            is_formula = isinstance(cell.value, str) and cell.value.startswith("=")
            try:
                value = _safe_value(cell.value)
            except IllegalCharacterError:
                value = "<invalid>"
            row_cells.append({"ref": cell.coordinate, "value": value, "isFormula": is_formula})
        rows_out.append(row_cells)

    result = {
        "sheet": sheet_name,
        "columns": [get_column_letter(c) for c in range(1, n_cols + 1)],
        "rows": rows_out,
        "totalRows": ws.max_row or 0,
        "totalCols": ws.max_column or 0,
    }
    wb.close()
    return result
