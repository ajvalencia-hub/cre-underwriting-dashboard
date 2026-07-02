"""Raw grid extraction from Excel/CSV rent rolls and T-12s: sheet listing,
best-guess header row detection, merged-cell fill-through, and numeric
parsing that handles $, commas, and parentheses-as-negative — the messy
formatting real rent rolls and operating statements actually use.
"""

import csv
import re
from pathlib import Path

import openpyxl

_NUMERIC_RE = re.compile(r"[^0-9.\-]")


def parse_numeric(value) -> float | None:
    """'$1,234.50' -> 1234.5, '(1,234)' -> -1234.0, '-' -> None, 12.3 -> 12.3."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text in ("", "-", "—", "N/A", "n/a"):
        return None
    negative = text.startswith("(") and text.endswith(")")
    if negative:
        text = text[1:-1]
    text = text.replace("%", "")
    cleaned = _NUMERIC_RE.sub("", text)
    if cleaned in ("", "-", "."):
        return None
    try:
        num = float(cleaned)
    except ValueError:
        return None
    return -num if negative else num


def list_sheets(path: Path, ext: str) -> list[str]:
    if ext == "csv":
        return ["Sheet1"]
    wb = openpyxl.load_workbook(path, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def _grid_from_csv(path: Path) -> list[list]:
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        return [row for row in csv.reader(f)]


def _grid_from_sheet(path: Path, sheet_name: str) -> list[list]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.worksheets[0]

    grid = [[cell.value for cell in row] for row in ws.iter_rows()]

    # Fill merged cell ranges with their top-left value so downstream code
    # sees a fully populated grid instead of blanks under a merge.
    for merged_range in ws.merged_cells.ranges:
        top_left_value = grid[merged_range.min_row - 1][merged_range.min_col - 1]
        for r in range(merged_range.min_row - 1, merged_range.max_row):
            for c in range(merged_range.min_col - 1, merged_range.max_col):
                if r < len(grid) and c < len(grid[r]):
                    grid[r][c] = top_left_value

    wb.close()
    return grid


def _guess_header_row(grid: list[list], max_scan_rows: int = 8) -> int:
    """Score the first few rows by how many non-empty, non-numeric-looking
    cells they have (headers are mostly text; data rows are mostly numbers).
    """
    best_idx, best_score = 0, -1
    for i, row in enumerate(grid[:max_scan_rows]):
        non_empty = [c for c in row if c is not None and str(c).strip() != ""]
        text_like = [c for c in non_empty if parse_numeric(c) is None]
        score = len(text_like)
        if score > best_score:
            best_score, best_idx = score, i
    return best_idx


def extract_grid(path: Path, ext: str, sheet_name: str | None = None) -> dict:
    """Returns {"sheet", "headerRowIndex", "headers", "rows"} where "rows" is
    the list of data rows (as raw cell values) below the detected header.
    """
    grid = _grid_from_csv(path) if ext == "csv" else _grid_from_sheet(path, sheet_name or "")
    grid = [row for row in grid if any(c is not None and str(c).strip() != "" for c in row)]

    if not grid:
        return {"sheet": sheet_name or "Sheet1", "headerRowIndex": -1, "headers": [], "rows": []}

    header_idx = _guess_header_row(grid)
    headers = [str(c).strip() if c is not None else "" for c in grid[header_idx]]
    rows = grid[header_idx + 1 :]

    return {"sheet": sheet_name or "Sheet1", "headerRowIndex": header_idx, "headers": headers, "rows": rows}
