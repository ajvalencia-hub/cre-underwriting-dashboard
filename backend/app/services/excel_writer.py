import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string


def _parse_cell_ref(ref: str) -> tuple[str | None, str]:
    """'Assumptions!C7' -> ('Assumptions', 'C7'); 'C7' -> (None, 'C7')."""
    if "!" in ref:
        sheet, cell = ref.split("!", 1)
        return sheet, cell
    return None, ref


def _lookup_defined_name(wb, ref: str):
    """Workbook-scoped names live in wb.defined_names; worksheet-scoped ones
    only in ws.defined_names (verified against openpyxl 3.1.5), listed by
    parse_workbook as 'Sheet!Name'. A defined name can never contain '!' but
    a sheet title can, so split on the last one.
    """
    if "!" in ref:
        sheet_title, name = ref.rsplit("!", 1)
        if sheet_title in wb.sheetnames:
            return wb[sheet_title].defined_names.get(name)
        return None
    return wb.defined_names.get(ref)


def _resolve_scalar_cell(wb, entry: dict):
    """Returns (worksheet, coord) or None. coord may still be a multi-cell
    range (e.g. 'B1:C2') when a named range spans one — callers MUST check for
    ':' before treating it as a single cell, since ws['B1:C2'] returns a tuple
    of rows, not a Cell.
    """
    if entry["target"] == "namedRange":
        defined_name = _lookup_defined_name(wb, entry["ref"])
        if defined_name is None:
            return None
        destinations = list(defined_name.destinations)
        if not destinations:
            return None
        sheet_title, coord = destinations[0]
        return wb[sheet_title], coord.replace("$", "")

    sheet_name, coord = _parse_cell_ref(entry["ref"])
    sheet_name = sheet_name or entry.get("sheet")
    if sheet_name is None or sheet_name not in wb.sheetnames:
        return None
    return wb[sheet_name], coord


def _coerce_value(value: Any, cell_type: str | None = None) -> Any:
    if isinstance(value, str) and cell_type == "date":
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError:
            return value
    return value


def _is_formula_cell(cell) -> bool:
    return isinstance(cell.value, str) and cell.value.startswith("=")


def _merge_anchor(ws, cell):
    """A MergedCell (any cell of a merged range except its top-left) is
    read-only — writing to it raises AttributeError. Mapping UIs display a
    merged range as one cell, so the user's intent is the anchor: remap there.
    """
    if not isinstance(cell, MergedCell):
        return cell
    for rng in ws.merged_cells.ranges:
        if cell.coordinate in rng:
            return ws.cell(row=rng.min_row, column=rng.min_col)
    return cell


def inject_values(
    template_path: Path, output_path: Path, mappings: dict, values: dict
) -> dict:
    """Copy the template and write mapped values into it, leaving everything else untouched.

    openpyxl writes raw values but never recalculates formulas, so downstream
    consumers must either open the file in Excel (which recalcs automatically because
    we set fullCalcOnLoad) or run the optional LibreOffice headless recalc pass.
    """
    shutil.copyfile(template_path, output_path)

    keep_vba = template_path.suffix.lower() == ".xlsm"
    wb = openpyxl.load_workbook(output_path, keep_vba=keep_vba)

    written: list[str] = []
    warnings: list[str] = []

    for field_id, entry in mappings.items():
        if field_id not in values or values[field_id] in (None, "", []):
            continue
        value = values[field_id]

        if entry["target"] in ("namedRange", "cell"):
            resolved = _resolve_scalar_cell(wb, entry)
            if resolved is None:
                warnings.append(f"Could not resolve mapped cell for '{field_id}' — skipped")
                continue
            ws, coord = resolved
            if ":" in coord:
                warnings.append(
                    f"'{field_id}' maps to the multi-cell range {ws.title}!{coord} — "
                    "value NOT written; map it to a single cell instead"
                )
                continue
            cell = _merge_anchor(ws, ws[coord])
            if _is_formula_cell(cell):
                warnings.append(
                    f"'{field_id}' maps to a formula cell {ws.title}!{coord} — "
                    "value NOT written to avoid breaking the model"
                )
                continue
            cell.value = _coerce_value(value)
            written.append(field_id)

        elif entry["target"] == "table":
            sheet_name = entry.get("sheet")
            if sheet_name is None or sheet_name not in wb.sheetnames:
                warnings.append(f"Could not resolve mapped sheet for '{field_id}' — skipped")
                continue
            ws = wb[sheet_name]

            col_letter, start_row = coordinate_from_string(entry["anchor"])
            start_col_idx = column_index_from_string(col_letter)
            column_order = entry.get("columnOrder") or ["key", "value"]

            rows = value if isinstance(value, list) else []
            skipped_formula_cells = 0
            skipped_merged_cells = 0
            for r_offset, row in enumerate(rows):
                for c_offset, col_id in enumerate(column_order):
                    cell = ws.cell(row=start_row + r_offset, column=start_col_idx + c_offset)
                    if isinstance(cell, MergedCell):
                        # Writing to the anchor here could clobber layout the
                        # template author merged deliberately — skip and count.
                        skipped_merged_cells += 1
                        continue
                    if _is_formula_cell(cell):
                        skipped_formula_cells += 1
                        continue
                    cell.value = _coerce_value(row.get(col_id))
            if skipped_formula_cells:
                warnings.append(
                    f"'{field_id}' skipped {skipped_formula_cells} cell(s) that contained formulas"
                )
            if skipped_merged_cells:
                warnings.append(
                    f"'{field_id}' skipped {skipped_merged_cells} cell(s) inside merged ranges"
                )
            written.append(field_id)

    wb.calculation.fullCalcOnLoad = True
    wb.save(output_path)
    return {"written": written, "warnings": warnings}


def read_output_values(path: Path, mappings: dict, output_field_ids: list[str]) -> dict[str, Any]:
    """Read back computed output cells. Only meaningful after a real recalc pass
    (e.g. LibreOffice headless) has run against `path` — otherwise the cached
    formula results are stale (pre-edit) or missing entirely, since openpyxl never
    evaluates formulas itself.
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    results: dict[str, Any] = {}
    for field_id in output_field_ids:
        entry = mappings.get(field_id)
        if entry is None or entry.get("target") not in ("namedRange", "cell"):
            continue
        resolved = _resolve_scalar_cell(wb, entry)
        if resolved is None:
            continue
        ws, coord = resolved
        if ":" in coord:
            continue
        value = _merge_anchor(ws, ws[coord]).value
        if value is not None:
            results[field_id] = value

    wb.close()
    return results
